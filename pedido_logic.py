"""
Lógica de negocio para el cálculo del pedido semanal de helado.

Flujo: ventas (cajas terminadas + mixventas) + stock → pedido → carrito Excel
"""

import base64
import logging
import math
import os
import re
from datetime import date, timedelta
from io import BytesIO
from pathlib import Path

import pandas as pd
import requests
from openpyxl import load_workbook
from rapidfuzz import fuzz, process

logger = logging.getLogger(__name__)

DATA_DIR = Path(__file__).parent / "data"
MAPEO_PATH = Path(__file__).parent / "mapeo_productos.csv"
PLAN_PATH = DATA_DIR / "compras_semanales_actual.csv"

# Ruta relativa al root del repo (para la API de contenidos de GitHub, que no
# acepta rutas absolutas de filesystem). Debe coincidir con DATA_DIR / archivo.
TEMPLATE_REPO_PATH = "data/carrito_template.xlsx"

# Repo por defecto para la persistencia vía GitHub (ver actualizar_archivo_en_github).
# Un solo repo válido para esta app (ver AGENTS.md §2) — no hace falta configurarlo
# por secret salvo que se quiera apuntar a otro lado.
GITHUB_REPO_DEFAULT = "ayankilevich-cpu/pedido_helado"


class ArchivoInvalidoError(Exception):
    """El archivo subido no tiene el formato esperado por un loader de negocio."""


# ---------------------------------------------------------------------------
# 1. Carga de archivos de ventas
# ---------------------------------------------------------------------------

def cargar_cajas_terminadas(file) -> pd.DataFrame:
    """Lee el .xls de cajas terminadas y agrupa por sabor (consolidado)."""
    try:
        df = pd.read_excel(file, engine="xlrd")
    except Exception as e:
        raise ArchivoInvalidoError(
            f"No se pudo leer el archivo de Cajas Terminadas como .xls: {e}"
        ) from e

    faltantes = {"artdescrip", "ctecantidad"} - set(df.columns)
    if faltantes:
        raise ArchivoInvalidoError(
            "El archivo de Cajas Terminadas no tiene las columnas esperadas "
            f"({', '.join(sorted(faltantes))}). ¿Es el archivo correcto para este cuadro?"
        )

    agrupado = (
        df.groupby("artdescrip", as_index=False)["ctecantidad"]
        .sum()
        .rename(columns={"artdescrip": "nombre_venta", "ctecantidad": "venta"})
    )
    agrupado["tipo"] = "granel"
    return agrupado


def cargar_mixventas(file) -> pd.DataFrame:
    """Lee el .xls de mixventas; filtra productos con bultos > 0 y subgrupo == 1.

    Productos granel vendidos a mayoristas (ej. baldes 7,8 kg) se detectan
    automáticamente por patrón en el nombre y se etiquetan como "granel"
    con el nombre en MAYÚSCULAS para coincidir con cajas terminadas.
    """
    try:
        df = pd.read_excel(file, engine="xlrd")
    except Exception as e:
        raise ArchivoInvalidoError(
            f"No se pudo leer el archivo de Mix Ventas como .xls: {e}"
        ) from e

    faltantes = {"artdescrip", "subgrupo", "bultos"} - set(df.columns)
    if faltantes:
        raise ArchivoInvalidoError(
            "El archivo de Mix Ventas no tiene las columnas esperadas "
            f"({', '.join(sorted(faltantes))}). ¿Es el archivo correcto para este cuadro?"
        )

    mask = (df["subgrupo"] == 1) & (df["bultos"].notna()) & (df["bultos"] > 0)
    resultado = (
        df.loc[mask, ["artdescrip", "bultos"]]
        .rename(columns={"artdescrip": "nombre_venta", "bultos": "venta"})
        .copy()
    )

    _GRANEL_RE = r"\(7[,.]8"
    is_granel = resultado["nombre_venta"].str.contains(
        _GRANEL_RE, case=False, na=False
    )
    resultado["tipo"] = "empaquetado"
    resultado.loc[is_granel, "tipo"] = "granel"
    resultado.loc[is_granel, "nombre_venta"] = (
        resultado.loc[is_granel, "nombre_venta"].str.upper()
    )
    return resultado


# ---------------------------------------------------------------------------
# 2. Carga del archivo de stock
# ---------------------------------------------------------------------------

def cargar_stock(file) -> pd.DataFrame:
    """
    Lee el archivo de stock (.xlsx).
    Retorna DataFrame con: codigo, descripcion, stock_seg, stock_real.
    Fila 0 del Excel es el header.
    """
    try:
        df = pd.read_excel(file, header=None, skiprows=1)
    except Exception as e:
        raise ArchivoInvalidoError(
            f"No se pudo leer el archivo de Stock como .xlsx: {e}"
        ) from e

    if df.shape[1] < 7:
        raise ArchivoInvalidoError(
            "El archivo de Stock no tiene el formato esperado (se esperaban al "
            f"menos 7 columnas y tiene {df.shape[1]}). ¿Es el archivo correcto "
            "para este cuadro?"
        )
    if df.empty:
        raise ArchivoInvalidoError("El archivo de Stock no tiene filas de datos.")

    stock = pd.DataFrame({
        "codigo": df[1].astype(str).str.replace(r"\.0$", "", regex=True).str.strip(),
        "descripcion": df[4].astype(str).str.strip(),
        "grupo": df[0].astype(str).str.strip(),
        "stock_seg": pd.to_numeric(df[5], errors="coerce").fillna(0),
        "stock_real": pd.to_numeric(df[6], errors="coerce").fillna(0),
    })
    return stock


# ---------------------------------------------------------------------------
# 3. Mapeo de nombres de ventas → código del carrito
# ---------------------------------------------------------------------------

GRUPOS_GRANEL = {
    "H13- SABORES AL AGUA", "H14- SABORES COMUNES",
    "H15- SABORES ESPECIALES", "H16- SABORES PREMIUM",
}

GRUPOS_EMPAQUETADO = {
    "H17- PALITOS", "H18- IMPULSIVOS", "H21- TORTAS Y POSTRES",
    "H22- FAMILIAR", "H23- POTE 1 LTS", "H02- FRIZZIO",
    "H12- FRUTAS BAÑADAS", "H03- SUNDAE GO",
    "H10- CONGELADOS - FUDY", "H24- CONGELADOS - EASY FRUT",
    "H20- PRODUCTOS SIN TACC",
}

# Mapeos manuales conocidos que el fuzzy matching no resuelve bien
_OVERRIDE_GRANEL = {
    "MASCARPONE CON FRUTOS DEL BOSQUE (7,8KG)": "4000058",
    "MANGO AL AGUA 7.8KG": "4000907",
    "FRUTILLA (7,8KG)": "4000194",
    "MARACUYA (7,8KG)": "4000038",
    "DULCE DE LECHE CON NUEZ (7,8KG)": "4000068",
    "CHOCOLATE CON ALMENDRAS (7,8KG)": "4000067",
}

# Descripciones para productos que están en el carrito pero no en el archivo de stock
_DESC_EXTRA = {
    "4000907": "MANGO AL AGUA 7,800 KG GRIDO",
}

_OVERRIDE_EMPAQUETADO = {
    "Familiar Nº 1": "4000163",
    "Familiar Nº 2": "4000164",
    "Familiar Nº 3": "4000165",
    "Familiar Nº 4": "4000166",
    "Tentacion 1 Lt Chocolate": "4000158",
    "Tentacion 1 Lt Chocolate Con Almendras": "4000155",
    "Tentacion 1 Lt Cookie": "4000157",
    "Tentacion 1 Lt Crema Americana": "4000162",
    "Tentacion 1 Lt Dulce de Leche": "4000159",
    "Tentacion 1 Lt Dulce de Leche Granizado": "4000153",
    "Tentacion 1 Lt Frutilla": "4000160",
    "Tentacion 1 Lt Granizado": "4000152",
    "Tentacion 1 Lt Limon": "4000156",
    "Tentacion 1 Lt Vainilla": "4000161",
    "Tentación 1l Menta Granizada": "4000670",
    "Tentacion Toddy Galletitas": "4000341",
    "Cups Black x 3": "6002071",
    "Cups Conito x 5 Un.": "6000014",
    "Cups Familiar  x 3 Cucuruchos": "6003296",
    "Cups Vasitos x 6": "6000196",
    "Bastoncito de Muzarella": "6002672",
    "Pechuguita de Pollo": "6002673",
    "Mini Frizzio": "6002674",
    "Pizza Frizzio Mozzarella": "6002671",
    "Pizza Frizzio Integral": "6002675",
    "Pizza Moz. y Jamon Frizzio": "6002679",
    "Pizza Tipo Casera": "6002680",
    "Empanada de Carne Frizzio 4u x 80g": "6003299",
    "Empanada de Jyq Frizzio 4u x 80g": "6003300",
    "Casatta en Caja x 8": "4000142",
    "Casatta x Unidad": "4000142",
    "Grido Tops Mini Rocklets": "6000999",
    "Grido Tops x 204 U": "",
    "Lunchera Oficial Selección Argentina": "",
    "Pote Reutilizable": "6002422",
    "Salsas x 500 Grs": "6000430",
    "Mermelada Frutilla Grido": "6000973",
    "Palito Bombon en Caja x 20": "4000859",
    "Palito Bombón X10 Un.": "4000859",
    "Cremoso Americana en Caja x 20": "4000854",
    "Cremoso Frutilla en Caja x 20": "4000855",
    "Palito Cremoso Americana X10 U.": "4000854",
    "Palito Cremoso Americana\xa0X10 U.": "4000854",
    "Palito Cremoso Frutilla X10 U.": "4000855",
    "Palito Frutal Frutilla en Caja x 20": "4000138",
    "Palito Frutal Limon en Caja x 20": "4000139",
    "Palito Frutal Naranja en Caja X20": "4000141",
    "Palito Frutal Frutilla en Caja x 10": "4000856",
    "Palito Frutal Naranja en Caja X10": "4000858",
    "Almendrado en Caja x 8": "4000143",
    "Almendrado x Unidad": "4000143",
    "Bombon Crocante en Caja x 8": "4000145",
    "Bombon Crocante x Unidad": "4000145",
    "Bombon Escoces en Caja x 8": "4000147",
    "Bombon Escoces x Unidad": "4000147",
    "Bombon Frutezza X8u.": "4000867",
    "Bombon Frutezza Xu.": "4000867",
    "Bombon Suizo en Caja x 8": "4000146",
    "Bombon Suizo x Unidad": "4000146",
    "Crocantino": "4000144",
    "Barra Delicia Chocolate y Mani": "4000593",
    "Postre Veg. Maní 80 Gr x Un": "6002857",
    "Alfajor Secreto Cookies And Cream x Un.": "4000823",
    "Alfajor Secreto Cookies And Cream X6": "4000823",
    "Torta Grido Rellena": "4000446",
    "Torta Cookies And Cream": "4000835",
    "Torta Frutillas Con Crema": "4000892",
    "Frambuesas Doble Chocolate X120": "6002434",
    "Frutillas Doble Chocolate X120": "6002435",
}

_ALL_OVERRIDES = {**_OVERRIDE_GRANEL, **_OVERRIDE_EMPAQUETADO}

# Also index by normalized keys (replace \xa0 with space, strip)
_OVERRIDES_NORMALIZED = {}
for k, v in _ALL_OVERRIDES.items():
    norm_key = k.replace("\xa0", " ").strip()
    _OVERRIDES_NORMALIZED[norm_key] = v


def _normalizar(texto: str) -> str:
    """Normaliza texto para mejorar el fuzzy matching."""
    import re
    t = texto.upper().strip()
    t = re.sub(r"\(7[,.]8\s*KG?\)", "", t)
    t = re.sub(r"7[,.]800?\s*KG", "", t)
    t = re.sub(r"7[.]8\s*KG", "", t)
    for word in ("GRIDO", "PACK", "CAJAS", "CAJ.", " - "):
        t = t.replace(word, " ")
    t = re.sub(r"\s+", " ", t).strip()
    return t


def generar_mapeo(
    ventas: pd.DataFrame,
    stock_df: pd.DataFrame,
    mapeo_path: Path | str = MAPEO_PATH,
    score_threshold: int = 40,
) -> pd.DataFrame:
    """
    Si mapeo_productos.csv existe, lo carga y solo agrega productos nuevos.
    Si no existe, genera uno completo con overrides manuales + fuzzy matching
    restringido por grupo de productos.
    """
    mapeo_path = Path(mapeo_path)

    existente = None
    if mapeo_path.exists():
        existente = pd.read_csv(mapeo_path, dtype={"codigo_carrito": str})
        existente["codigo_carrito"] = (
            existente["codigo_carrito"]
            .astype(str)
            .str.replace(r"\.0$", "", regex=True)
            .str.strip()
        )
        nombres_ya = set(existente["nombre_venta"].str.strip())
        nuevos = ventas[~ventas["nombre_venta"].str.strip().isin(nombres_ya)]
        if nuevos.empty:
            return existente
        ventas_a_mapear = nuevos
    else:
        ventas_a_mapear = ventas

    codigo_a_desc = dict(zip(stock_df["codigo"], stock_df["descripcion"]))
    codigo_a_desc.update(_DESC_EXTRA)

    granel_mask = stock_df["grupo"].isin(GRUPOS_GRANEL)
    empaq_mask = stock_df["grupo"].isin(GRUPOS_EMPAQUETADO)
    subsets = {
        "granel": stock_df[granel_mask].reset_index(drop=True),
        "empaquetado": stock_df[empaq_mask].reset_index(drop=True),
    }

    filas = []
    for _, row in ventas_a_mapear.iterrows():
        nombre = row["nombre_venta"]
        tipo = row["tipo"]

        nombre_clean = nombre.replace("\xa0", " ").strip()
        override_cod = _OVERRIDES_NORMALIZED.get(nombre_clean)
        if override_cod is not None:
            cod = override_cod
            if cod:
                filas.append({
                    "nombre_venta": nombre,
                    "tipo": tipo,
                    "codigo_carrito": cod,
                    "descripcion_carrito": codigo_a_desc.get(cod, "???"),
                    "score": 100,
                })
            else:
                filas.append({
                    "nombre_venta": nombre,
                    "tipo": tipo,
                    "codigo_carrito": "",
                    "descripcion_carrito": "NO APLICA",
                    "score": 0,
                })
            continue

        subset = subsets.get(tipo, stock_df)
        descs = subset["descripcion"].tolist()
        cods = subset["codigo"].tolist()
        descs_norm = [_normalizar(d) for d in descs]
        nombre_norm = _normalizar(nombre)

        resultado = process.extractOne(
            nombre_norm,
            descs_norm,
            scorer=fuzz.token_sort_ratio,
            score_cutoff=score_threshold,
        )
        if resultado:
            match_text, score, idx = resultado
            filas.append({
                "nombre_venta": nombre,
                "tipo": tipo,
                "codigo_carrito": cods[idx],
                "descripcion_carrito": descs[idx],
                "score": int(score),
            })
        else:
            filas.append({
                "nombre_venta": nombre,
                "tipo": tipo,
                "codigo_carrito": "",
                "descripcion_carrito": "SIN MAPEO",
                "score": 0,
            })

    nuevos_df = pd.DataFrame(filas)

    if existente is not None:
        resultado_final = pd.concat([existente, nuevos_df], ignore_index=True)
    else:
        resultado_final = nuevos_df

    try:
        resultado_final.to_csv(mapeo_path, index=False)
    except OSError as e:
        logger.warning("No se pudo guardar mapeo_productos.csv en %s: %s", mapeo_path, e)
    return resultado_final


def cargar_mapeo_disco(mapeo_path: Path | str = MAPEO_PATH) -> pd.DataFrame | None:
    """Carga el mapeo existente desde disco, si hay uno.

    No depende de Streamlit: la resolución contra `session_state` (mapeo
    editado en la sesión actual, aún no persistido) vive en `app_pedido.py`.
    """
    mapeo_path = Path(mapeo_path)
    if mapeo_path.exists():
        df = pd.read_csv(mapeo_path, dtype={"codigo_carrito": str})
        df["codigo_carrito"] = (
            df["codigo_carrito"]
            .astype(str)
            .str.replace(r"\.0$", "", regex=True)
            .str.strip()
        )
        return df
    return None


# ---------------------------------------------------------------------------
# 4. Planificación semanal (compras planificadas por semana)
# ---------------------------------------------------------------------------

# Códigos de ajuste devueltos en la columna `ajuste_plan` del pedido
AJUSTE_NINGUNO = ""
AJUSTE_SUBE = "up"          # cálculo < 97% del plan → se sube al mínimo
AJUSTE_BAJA = "down"        # cálculo > 105% del plan → se baja al máximo
AJUSTE_PLAN_CERO = "zero"   # producto en plan con valor 0
AJUSTE_SIN_PLAN = "missing" # producto sin entrada en el plan

_MESES_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}

_RE_SEMANA = re.compile(r"^Semana_(\d+)_([A-Za-zñÑáéíóúÁÉÍÓÚ]+)_(\d{4})$")


def cargar_planificacion(file) -> pd.DataFrame:
    """Lee el CSV de planificación semanal de compras.

    Espera columnas: codigo_homologado, descripcion, linea_producto, categoria,
    Semana_<n>_<Mes>_<Año>... y total_temporada.

    Normaliza codigo_homologado a string sin sufijos `.0` ni espacios.
    """
    if hasattr(file, "seek"):
        try:
            file.seek(0)
        except OSError:
            pass
    df = pd.read_csv(file, dtype={"codigo_homologado": str})
    df["codigo_homologado"] = (
        df["codigo_homologado"]
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
    )
    return df


def _guardar_planificacion(file, dest: Path | str = PLAN_PATH) -> tuple[bytes | None, bool]:
    """Guarda copia local del CSV de planificación.

    Devuelve (contenido, guardado_en_disco). `guardado_en_disco=False` en
    Streamlit Cloud significa que el archivo NO sobrevive a un reinicio del
    contenedor (filesystem efímero, ver AGENTS.md §7) — el caller debe avisar.
    """
    dest = Path(dest)
    if hasattr(file, "read"):
        if hasattr(file, "seek"):
            try:
                file.seek(0)
            except OSError:
                pass
        contenido = file.read()
        if hasattr(file, "seek"):
            try:
                file.seek(0)
            except OSError:
                pass
    else:
        contenido = Path(file).read_bytes()
    try:
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(contenido)
        guardado = True
    except OSError as e:
        logger.warning("No se pudo guardar la planificación en %s: %s", dest, e)
        guardado = False
    return contenido, guardado


def obtener_planificacion() -> Path | None:
    """Retorna la ruta de la última planificación guardada, si existe."""
    return PLAN_PATH if PLAN_PATH.exists() else None


def _parsear_semana(col: str) -> tuple[int, int, int] | None:
    """De `Semana_18_Mayo_2026` → (año, semana, mes_hint). Retorna None si no matchea."""
    m = _RE_SEMANA.match(col)
    if not m:
        return None
    semana = int(m.group(1))
    mes_nombre = m.group(2).lower()
    año = int(m.group(3))
    mes = _MESES_ES.get(mes_nombre, 0)
    return año, semana, mes


def _primer_domingo_del_año(año: int) -> date:
    """Primer domingo del año (1/1 si cae domingo, si no el siguiente domingo)."""
    d = date(año, 1, 1)
    return d + timedelta(days=(6 - d.weekday()) % 7)


def inicio_semana(año: int, semana: int) -> date:
    """
    Domingo de inicio de la `Semana N / Año`.
    Convención del CSV: las semanas van domingo→sábado y la numeración se reinicia
    cada año en la Sem 1 = primer domingo del año.
    """
    return _primer_domingo_del_año(año) + timedelta(days=(semana - 1) * 7)


def fin_semana(año: int, semana: int) -> date:
    """Sábado de cierre de la `Semana N / Año`."""
    return inicio_semana(año, semana) + timedelta(days=6)


def obtener_semanas(plan_df: pd.DataFrame) -> list[tuple[str, date]]:
    """Devuelve [(nombre_columna, fecha_inicio_domingo)] de las columnas semanales del CSV."""
    out: list[tuple[str, date]] = []
    for col in plan_df.columns:
        parsed = _parsear_semana(col)
        if parsed is None:
            continue
        año, semana, _ = parsed
        try:
            fecha = inicio_semana(año, semana)
        except ValueError:
            continue
        out.append((col, fecha))
    out.sort(key=lambda x: x[1])
    return out


def semana_default(plan_df: pd.DataFrame, hoy: date | None = None) -> str | None:
    """Columna semanal cuyo rango [domingo, sábado] contiene `hoy`. Si hoy queda fuera del CSV, la más cercana."""
    semanas = obtener_semanas(plan_df)
    if not semanas:
        return None
    hoy = hoy or date.today()
    for col, inicio in semanas:
        if inicio <= hoy <= inicio + timedelta(days=6):
            return col
    if hoy < semanas[0][1]:
        return semanas[0][0]
    return semanas[-1][0]


# ---------------------------------------------------------------------------
# 5. Cálculo del pedido
# ---------------------------------------------------------------------------

def calcular_pedido(
    ventas: pd.DataFrame,
    mapeo: pd.DataFrame,
    stock_df: pd.DataFrame,
    pct_stock_seg: int = 100,
    pct_ajuste_venta: float = 0,
    plan_df: pd.DataFrame | None = None,
    plan_col: str | None = None,
    pct_plan_min: float = 3.0,
    pct_plan_max: float = 5.0,
    modo_replicar_venta: bool = False,
) -> pd.DataFrame:
    """
    Calcula el pedido de reposición por producto.

    Lógica base (modo normal):
    - venta_ajustada = venta * (1 + pct_ajuste_venta/100)
    - pedido_calc = max(0, ceil(venta_ajustada + stock_seg * pct_stock_seg/100 - stock_real))

    Modo replicar venta (modo_replicar_venta=True):
    - pedido = ceil(venta * (1 + pct_ajuste_venta/100))
    - Ignora stock real, stock de seguridad y planificación.
    - Útil para semanas donde se quiere reponer exactamente lo vendido.

    Ajuste por planificación semanal (solo en modo normal):
    - lim_min = ceil(plan * (1 - pct_plan_min/100))
    - lim_max = floor(plan * (1 + pct_plan_max/100))
    - Si plan > 0:
        pedido = clip(pedido_calc, lim_min, lim_max)
    - Si plan == 0 o el producto no está en el plan:
        pedido = 0 y se marca para revisión manual.

    Retorna DataFrame con columnas:
        codigo_carrito, descripcion, grupo, venta, stock_real, stock_seg,
        plan_sem, pedido_calc, ajuste_plan, pedido
    Las tres columnas relacionadas con el plan estarán siempre presentes:
    si no hay plan cargado, plan_sem=NaN, pedido_calc=pedido, ajuste_plan="".
    """
    mapeo_valido = mapeo[
        (mapeo["codigo_carrito"].notna())
        & (mapeo["codigo_carrito"].astype(str).str.strip() != "")
        & (mapeo["descripcion_carrito"] != "SIN MAPEO")
        & (mapeo["descripcion_carrito"] != "NO APLICA")
    ].copy()
    mapeo_valido["codigo_carrito"] = (
        mapeo_valido["codigo_carrito"]
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
    )

    ventas_con_codigo = ventas.merge(
        mapeo_valido[["nombre_venta", "codigo_carrito"]],
        on="nombre_venta",
        how="inner",
    )

    venta_por_codigo = (
        ventas_con_codigo
        .groupby("codigo_carrito", as_index=False)["venta"]
        .sum()
    )

    stock_df = stock_df.copy()
    stock_df["codigo"] = stock_df["codigo"].astype(str)

    # ── SKUs granel sin venta ──────────────────────────────────────────────────
    # Si un sabor granel no tuvo movimiento (faltante, quiebre de stock, baja
    # estacional), igual aparece en el listado con venta=0 y pedido=0 editable.
    # Empaquetados se excluyen: su ausencia puede ser discontinuación o no pedido.
    codigos_con_venta = set(venta_por_codigo["codigo_carrito"].astype(str))
    sin_venta = stock_df[
        stock_df["grupo"].isin(GRUPOS_GRANEL | GRUPOS_EMPAQUETADO)
        & ~stock_df["codigo"].isin(codigos_con_venta)
    ][["codigo"]].copy()
    sin_venta = sin_venta.rename(columns={"codigo": "codigo_carrito"})
    sin_venta["venta"] = 0.0
    sin_venta["_sin_venta"] = True  # marca para forzar pedido=0 al final

    venta_por_codigo = pd.concat(
        [venta_por_codigo, sin_venta],
        ignore_index=True,
    )
    # ── Fin SKUs sin venta ────────────────────────────────────────────────────

    pedido_df = venta_por_codigo.merge(
        stock_df[["codigo", "descripcion", "grupo", "stock_seg", "stock_real"]],
        left_on="codigo_carrito",
        right_on="codigo",
        how="left",
    )

    pedido_df["stock_seg"] = pedido_df["stock_seg"].fillna(0)
    pedido_df["stock_real"] = pedido_df["stock_real"].fillna(0)
    if "_sin_venta" not in pedido_df.columns:
        pedido_df["_sin_venta"] = False
    pedido_df["_sin_venta"] = pedido_df["_sin_venta"].fillna(False)

    # Ajuste de venta por estacionalidad: venta * (1 + pct/100)
    pedido_df["venta"] = pedido_df["venta"] * (1 + pct_ajuste_venta / 100)

    if pedido_df["descripcion"].isna().any():
        # mapeo_valido puede tener codigo_carrito duplicado (varios nombre_venta
        # apuntando al mismo código) — construir el dict directo desde ahí evita
        # desalinear las dos secuencias del zip al re-filtrar `mapeo` por isin().
        mapeo_desc = dict(zip(mapeo_valido["codigo_carrito"], mapeo_valido["descripcion_carrito"]))
        mask = pedido_df["descripcion"].isna()
        pedido_df.loc[mask, "descripcion"] = (
            pedido_df.loc[mask, "codigo_carrito"].map(mapeo_desc)
        )
        pedido_df["grupo"] = pedido_df["grupo"].fillna("")

    # ── Modo replicar venta ────────────────────────────────────────────────────
    # pedido = ceil(venta_ajustada), ignorando stock real, stock de seguridad
    # y planificación. El ajuste % de venta sí aplica (ya fue aplicado arriba).
    if modo_replicar_venta:
        pedido_df["pedido_calc"] = pedido_df["venta"].apply(
            lambda v: max(0, math.ceil(v))
        ).astype(int)
        pedido_df["plan_sem"] = pd.NA
        pedido_df["pedido"] = pedido_df["pedido_calc"].astype(int)
        pedido_df["ajuste_plan"] = AJUSTE_NINGUNO
        # Granel sin venta → pedido 0 editable (también en modo replicar)
        mask_sv = pedido_df["_sin_venta"] == True
        pedido_df.loc[mask_sv, "pedido"] = 0
        pedido_df.loc[mask_sv, "pedido_calc"] = 0
        pedido_df["pedido_inicial"] = pedido_df["pedido"].astype(int)
        cols = [
            "codigo_carrito", "descripcion", "grupo",
            "venta", "stock_real", "stock_seg",
            "plan_sem", "pedido_calc", "ajuste_plan",
            "pedido_inicial", "pedido",
        ]
        return pedido_df[cols].sort_values("grupo").reset_index(drop=True)
    # ── Fin modo replicar venta ───────────────────────────────────────────────

    factor = pct_stock_seg / 100

    def _calc_pedido(row):
        return max(0, math.ceil(row["venta"] + row["stock_seg"] * factor - row["stock_real"]))

    pedido_df["pedido_calc"] = pedido_df.apply(_calc_pedido, axis=1).astype(int)

    # Ajuste por planificación semanal
    aplicar_plan = plan_df is not None and plan_col is not None and plan_col in plan_df.columns
    if aplicar_plan:
        plan_lookup = plan_df.set_index("codigo_homologado")[plan_col]
        pedido_df["plan_sem"] = (
            pedido_df["codigo_carrito"].map(plan_lookup).astype(float)
        )

        def _ajustar(row):
            calc = int(row["pedido_calc"])
            plan = row["plan_sem"]
            if pd.isna(plan):
                return pd.Series({"pedido": 0, "ajuste_plan": AJUSTE_SIN_PLAN})
            plan_num = float(plan)
            if plan_num <= 0:
                return pd.Series({"pedido": 0, "ajuste_plan": AJUSTE_PLAN_CERO})
            lim_min = math.ceil(plan_num * (1 - pct_plan_min / 100))
            lim_max = math.floor(plan_num * (1 + pct_plan_max / 100))
            if lim_max < lim_min:
                lim_max = lim_min
            if calc < lim_min:
                return pd.Series({"pedido": int(lim_min), "ajuste_plan": AJUSTE_SUBE})
            if calc > lim_max:
                return pd.Series({"pedido": int(lim_max), "ajuste_plan": AJUSTE_BAJA})
            return pd.Series({"pedido": calc, "ajuste_plan": AJUSTE_NINGUNO})

        ajustado = pedido_df.apply(_ajustar, axis=1)
        pedido_df["pedido"] = ajustado["pedido"].astype(int)
        pedido_df["ajuste_plan"] = ajustado["ajuste_plan"].astype(str)
    else:
        pedido_df["plan_sem"] = pd.NA
        pedido_df["pedido"] = pedido_df["pedido_calc"].astype(int)
        pedido_df["ajuste_plan"] = AJUSTE_NINGUNO

    # Granel sin venta → pedido 0 editable (visible pero sin cantidad sugerida)
    mask_sv = pedido_df["_sin_venta"] == True
    pedido_df.loc[mask_sv, "pedido"] = 0
    pedido_df.loc[mask_sv, "pedido_calc"] = 0

    pedido_df["pedido_inicial"] = pedido_df["pedido"].astype(int)

    cols = [
        "codigo_carrito", "descripcion", "grupo",
        "venta", "stock_real", "stock_seg",
        "plan_sem", "pedido_calc", "ajuste_plan",
        "pedido_inicial", "pedido",
    ]
    return pedido_df[cols].sort_values("grupo").reset_index(drop=True)


# ---------------------------------------------------------------------------
# 6. Escritura del carrito Excel
# ---------------------------------------------------------------------------

def validar_plantilla_carrito(contenido: bytes) -> tuple[bool, str]:
    """
    Comprueba que el Excel sea un Modelo de Carrito usable (hoja tipo export Grido).
    Retorna (ok, mensaje_error). Si ok, mensaje_error es "".
    """
    if not contenido or len(contenido) < 100:
        return False, "El archivo está vacío o es demasiado pequeño."

    try:
        bio = BytesIO(contenido)
        wb = load_workbook(bio, read_only=True, data_only=True)
    except Exception as e:
        return False, f"No es un Excel válido (.xlsx): {e}"

    ok = True
    msg = ""
    try:
        ws = wb.active
        if ws.max_row < 2:
            ok, msg = False, "El Excel no tiene filas de datos (solo encabezado o vacío)."
        else:
            # Cabecera esperada del export: Codigo en col B, precio en col I (fila 1)
            h_b = ws.cell(row=1, column=2).value
            h_i = ws.cell(row=1, column=9).value
            cod_ok = h_b is not None and "codigo" in str(h_b).lower()
            pre_ok = h_i is not None and "precio" in str(h_i).lower()
            if not (cod_ok and pre_ok):
                ok, msg = (
                    False,
                    "No coincide con la plantilla del Modelo de Carrito: en la fila 1 "
                    "deben aparecer columnas tipo «Codigo» (B) y «precio» (I). "
                    "¿Subiste el archivo de Stock por error? Ese va en «Stock», no aquí.",
                )
            else:
                c2 = ws.cell(row=2, column=2).value
                if c2 is None or not str(c2).strip():
                    ok, msg = False, "No hay código de producto en la fila 2 (columna B)."
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return ok, msg


def _guardar_plantilla(file, dest: Path | str = DATA_DIR / "carrito_template.xlsx") -> tuple[bytes | None, bool]:
    """Guarda una copia local de la plantilla del carrito.

    Devuelve (contenido, guardado_en_disco). `guardado_en_disco=False` en
    Streamlit Cloud significa que el archivo NO sobrevive a un reinicio del
    contenedor (filesystem efímero, ver AGENTS.md §7) — el caller debe avisar.
    Para persistencia real entre reinicios ver `actualizar_archivo_en_github`.
    """
    dest = Path(dest)
    if hasattr(file, "read"):
        contenido = file.read()
        file.seek(0)
    else:
        contenido = Path(file).read_bytes()

    try:
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(contenido)
        guardado = True
    except OSError as e:
        logger.warning("No se pudo guardar la plantilla en %s: %s", dest, e)
        guardado = False
    return contenido, guardado


def obtener_plantilla() -> Path | None:
    """Retorna la ruta de la última plantilla guardada, si existe."""
    p = DATA_DIR / "carrito_template.xlsx"
    return p if p.exists() else None


# ---------------------------------------------------------------------------
# 6bis. Persistencia remota (GitHub) — filesystem efímero de Streamlit Cloud
# ---------------------------------------------------------------------------
#
# El disco de Streamlit Cloud no sobrevive a un reinicio del contenedor
# (sleep por inactividad, redeploy, etc.) — ver AGENTS.md §7. `_guardar_plantilla`
# deja la plantilla usable para la sesión actual, pero al próximo arranque en
# frío la app vuelve a leer lo que esté commiteado en el repo. Esta función
# automatiza el procedimiento manual ya documentado en AGENTS.md §11
# ("commit + push + reboot") vía la API de contenidos de GitHub, para que una
# plantilla subida por UI quede persistida de verdad sin depender de que
# alguien la vuelva a subir manualmente al repo.
#
# No depende de Streamlit: token y repo se resuelven en app_pedido.py desde
# st.secrets y se pasan como parámetros.

def actualizar_archivo_en_github(
    contenido: bytes,
    token: str,
    path_repo: str = TEMPLATE_REPO_PATH,
    repo: str = GITHUB_REPO_DEFAULT,
    mensaje: str = "chore: actualizar plantilla del carrito vía app",
    branch: str = "main",
) -> tuple[bool, str]:
    """Commitea `contenido` en `path_repo` del repo de GitHub (API de contenidos).

    Requiere un token con permiso de escritura sobre `repo` (fine-grained
    "Contents: Read and write" o classic con scope `repo`). Devuelve
    (ok, mensaje_error) — mensaje_error es "" si ok.
    """
    api_url = f"https://api.github.com/repos/{repo}/contents/{path_repo}"
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
    }
    try:
        resp_get = requests.get(api_url, headers=headers, params={"ref": branch}, timeout=15)
        sha_actual = resp_get.json().get("sha") if resp_get.status_code == 200 else None

        payload = {
            "message": mensaje,
            "content": base64.b64encode(contenido).decode("ascii"),
            "branch": branch,
        }
        if sha_actual:
            payload["sha"] = sha_actual

        resp_put = requests.put(api_url, headers=headers, json=payload, timeout=30)
    except requests.RequestException as e:
        logger.warning("Error de red al commitear %s a GitHub: %s", path_repo, e)
        return False, f"No se pudo conectar con GitHub: {e}"

    if resp_put.status_code in (200, 201):
        logger.info("Commit a GitHub OK: %s (%s)", path_repo, repo)
        return True, ""

    try:
        detalle = resp_put.json().get("message", resp_put.text)
    except ValueError:
        detalle = resp_put.text
    logger.warning(
        "Commit a GitHub falló (%s) para %s: %s", resp_put.status_code, path_repo, detalle
    )
    return False, f"GitHub respondió {resp_put.status_code}: {detalle}"


def cargar_datos_plantilla(
    template_path: str | Path,
) -> tuple[dict[str, float], dict[str, float], dict[str, float]]:
    """
    Extrae cubicaje (col F), peso (col G) y precio (col I) de la plantilla.
    Retorna ({codigo: cubicaje}, {codigo: precio}, {codigo: peso_kg}).

    Usa pandas para ser rápido (openpyxl en modo read_only iterando celdas tarda
    decenas de segundos en plantillas con cientos de filas y muchas columnas;
    pd.read_excel resuelve en ~1s gracias a la lectura vectorizada).
    """
    df = pd.read_excel(
        template_path,
        engine="openpyxl",
        usecols=[1, 5, 6, 8],  # B=Codigo, F=Cubicaje, G=Peso, I=precio (0-indexed)
        header=0,
        dtype={1: str},
    )
    df.columns = ["codigo", "cubicaje", "peso", "precio"]
    df = df.dropna(subset=["codigo"])
    df["codigo"] = df["codigo"].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    for c in ("cubicaje", "peso", "precio"):
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    cubicaje = dict(zip(df["codigo"], df["cubicaje"].astype(float)))
    pesos = dict(zip(df["codigo"], df["peso"].astype(float)))
    precios = dict(zip(df["codigo"], df["precio"].astype(float)))
    return cubicaje, precios, pesos


def cargar_cubicaje(template_path: str | Path) -> dict[str, float]:
    """Extrae el cubicaje (columna F) de la plantilla del carrito. Retorna {codigo: cubicaje}."""
    cubicaje, _, _ = cargar_datos_plantilla(template_path)
    return cubicaje


def escribir_carrito(
    template_path: str | Path,
    pedido_df: pd.DataFrame,
) -> BytesIO:
    """
    Abre la plantilla del carrito con openpyxl, escribe las cantidades
    en la columna C (índice 3 en openpyxl, fila 2+ porque fila 1 es header),
    y retorna un BytesIO con el archivo listo para descargar.
    """
    wb = load_workbook(template_path)
    ws = wb.active

    pedido_dict = dict(
        zip(
            pedido_df["codigo_carrito"].astype(str),
            pedido_df["pedido"],
        )
    )

    for row_idx in range(2, ws.max_row + 1):
        codigo_cell = ws.cell(row=row_idx, column=2).value
        if codigo_cell is None:
            continue
        codigo = str(codigo_cell).strip()
        if codigo in pedido_dict:
            ws.cell(row=row_idx, column=3).value = pedido_dict[codigo]

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
